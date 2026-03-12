import json
import os
import threading
import time
import sqlite3
import csv
import io
import secrets
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta
from functools import wraps
from dotenv import load_dotenv
from authlib.integrations.flask_client import OAuth
from werkzeug.security import generate_password_hash, check_password_hash
from flask import Flask, render_template, request, jsonify, Response, send_file, session, redirect, url_for
import takeout_runner
import takeout_downloader
import pyotp
from fpdf import FPDF
from openpyxl import Workbook
# ── Security additions ────────────────────────────────────────────────
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_wtf.csrf import CSRFProtect, CSRFError
from cryptography.fernet import Fernet
from cachetools import TTLCache

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv("AGENT_SECRET_KEY", secrets.token_hex(32))
if not app.secret_key:
    raise RuntimeError(
        "FLASK_SECRET_KEY is not set in .env! "
        "Generate one with: python -c \"import secrets; print(secrets.token_hex(32))\""
    )
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB upload limit
app.config['WTF_CSRF_SECRET_KEY'] = os.getenv('WTF_CSRF_SECRET_KEY', app.secret_key)

# ── Rate Limiter ──────────────────────────────────────────────────────
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=[],
    storage_uri="memory://"
)

# ── CSRF Protection ───────────────────────────────────────────────────
csrf = CSRFProtect(app)

# ── Fernet Encryption (for target account passwords) ─────────────────
_fernet_key = os.getenv('FERNET_KEY')
if not _fernet_key:
    # Instead of plain-text fallback, we now generate a session-persistent key if missing.
    # This prevents storing secrets in cleartext even if the user forgets to set a key.
    # Note: Passwords will be unreadable after server restart unless the key is saved to .env.
    _fernet_key = Fernet.generate_key().decode()
    print("\n" + "!"*60)
    print("  [!!!] CRITICAL SECURITY WARNING: FERNET_KEY not set in .env!")
    print("        A temporary key has been generated for this session.")
    print("        TARGET ACCOUNT PASSWORDS WILL BE LOST ON RESTART.")
    print(f"        To persist them, add this to your .env: FERNET_KEY={_fernet_key}")
    print("!"*60 + "\n")
_fernet = Fernet(_fernet_key.encode())

def encrypt_password(plain: str) -> str:
    """Encrypt a plaintext password for storage."""
    return _fernet.encrypt(plain.encode()).decode()

def decrypt_password(stored: str) -> str:
    """Decrypt a stored password. Falls back to returning stored value if decryption fails (migration)."""
    try:
        return _fernet.decrypt(stored.encode()).decode()
    except Exception:
        return stored  # Likely plaintext from before encryption was mandatory

# ── 413 handler ───────────────────────────────────────────────────────
@app.errorhandler(413)
def too_large(e):
    return jsonify({"error": "File too large. Maximum upload size is 16 MB."}), 413

# ── 429 handler ───────────────────────────────────────────────────────
@app.errorhandler(429)
def rate_limited(e):
    return render_template("login.html", error="Too many attempts. Please wait a minute and try again."), 429

# ── CSRF error handler ────────────────────────────────────────────────
@app.errorhandler(CSRFError)
def csrf_error(e):
    return jsonify({"error": "CSRF validation failed. Please refresh the page and try again."}), 400

# Authlib configuration
app.config['GOOGLE_CLIENT_ID'] = os.getenv("GOOGLE_CLIENT_ID")
app.config['GOOGLE_CLIENT_SECRET'] = os.getenv("GOOGLE_CLIENT_SECRET")

if not app.config['GOOGLE_CLIENT_ID']:
    print("WARNING: GOOGLE_CLIENT_ID is missing from your .env file! SSO will fail!")

oauth = OAuth(app)
google = oauth.register(
    name='google',
    server_metadata_url='https://accounts.google.com/.well-known/openid-configuration',
    client_kwargs={
        'scope': 'openid email profile'
    }
)

DB_FILE = "users.db"

def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS users (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT UNIQUE NOT NULL,
                    password_hash TEXT NOT NULL,
                    role TEXT NOT NULL,
                    mfa_secret TEXT,
                    mfa_enabled BOOLEAN DEFAULT 0,
                    settings TEXT DEFAULT '{}'
                 )''')
    c.execute('''CREATE TABLE IF NOT EXISTS audit_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                    username TEXT,
                    ip_address TEXT,
                    action TEXT,
                    details TEXT
                 )''')
    c.execute('''CREATE TABLE IF NOT EXISTS target_accounts (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    email TEXT NOT NULL,
                    password TEXT NOT NULL,
                    backup_codes TEXT,
                    create_status TEXT DEFAULT 'Pending',
                    download_status TEXT DEFAULT 'Pending',
                    owner TEXT NOT NULL DEFAULT 'admin',
                    UNIQUE(email, owner)
                 )''')
    c.execute('''CREATE TABLE IF NOT EXISTS invite_tokens (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    token TEXT UNIQUE NOT NULL,
                    email TEXT NOT NULL,
                    role TEXT NOT NULL DEFAULT 'user',
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    expires_at DATETIME NOT NULL,
                    used INTEGER DEFAULT 0
                 )''')
    c.execute('''CREATE TABLE IF NOT EXISTS agent_registrations (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    agent_id TEXT UNIQUE NOT NULL,
                    username TEXT NOT NULL,
                    token TEXT UNIQUE NOT NULL,
                    hostname TEXT,
                    last_seen DATETIME,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                 )''')
    c.execute('''CREATE TABLE IF NOT EXISTS agent_jobs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    agent_id TEXT NOT NULL,
                    email TEXT NOT NULL,
                    mode TEXT NOT NULL DEFAULT 'create',
                    status TEXT DEFAULT 'pending',
                    logs TEXT DEFAULT '',
                    settings TEXT DEFAULT '{}',
                    assigned_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    completed_at DATETIME
                 )''')
    c.execute('''CREATE TABLE IF NOT EXISTS agent_detailed_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    agent_id TEXT NOT NULL,
                    email TEXT NOT NULL,
                    log_line TEXT NOT NULL,
                    timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
                 )''')
    # Migrations for existing DBs — target_accounts and users tables
    for col, defval in [
        ("owner", "'admin'"),
        ("settings", "'{}'"),
    ]:
        try:
            # col is sanitized by the caller or matched against a whitelist
            allowed_cols = ["owner", "create_status", "download_status", "last_seen", "hostname", "settings"]
            if col not in allowed_cols:
                raise ValueError(f"Invalid column name: {col}")
            
            table = 'target_accounts' if col == 'owner' else 'users'
            # Bandit B608: Since table and col are whitelisted/controlled, this is safe
            # but we still prefer explicit mapping where possible.
            query = f"SELECT {col} FROM {table} LIMIT 1"
            c.execute(query)
        except sqlite3.OperationalError:
            c.execute(f"ALTER TABLE {'target_accounts' if col == 'owner' else 'users'} ADD COLUMN {col} TEXT NOT NULL DEFAULT {defval}")

    # Migration: add `settings` column to agent_jobs if missing (column added in security audit)
    try:
        c.execute("SELECT settings FROM agent_jobs LIMIT 1")
    except sqlite3.OperationalError:
        c.execute("ALTER TABLE agent_jobs ADD COLUMN settings TEXT NOT NULL DEFAULT '{}'")

    c.execute("SELECT * FROM users WHERE username = 'admin'")
    admin_row = c.fetchone()
    if not admin_row:
        c.execute("INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)",
                  ('admin', generate_password_hash('admin123'), 'admin'))
    else:
        # P2: Warn if default admin password still in use
        if check_password_hash(admin_row[2], 'admin123'):
            print("\n" + "="*60)
            print("  [!] SECURITY WARNING: Default admin password 'admin123' is")
            print("      still in use! Change it immediately via the admin panel.")
            print("="*60 + "\n")
    conn.commit()
    conn.close()


init_db()

def get_db_connection():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def send_email(to_email, subject, html_body):
    """Send an HTML email using SMTP credentials from .env."""
    smtp_server = os.getenv("SMTP_SERVER", "smtp.gmail.com")
    smtp_port = int(os.getenv("SMTP_PORT", 587))
    smtp_user = os.getenv("SMTP_USER", "")
    smtp_password = os.getenv("SMTP_PASSWORD", "")

    if not smtp_user or not smtp_password:
        print("WARNING: SMTP credentials not configured in .env")
        return False

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = f"Takeout Automator <{smtp_user}>"
    msg["To"] = to_email
    msg.attach(MIMEText(html_body, "html"))

    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.ehlo()
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.sendmail(smtp_user, to_email, msg.as_string())
        return True
    except Exception as e:
        print(f"Email send failed: {e}")
        return False

def log_activity(action, details=None, username=None):
    try:
        if not username:
            username = session.get('username', 'system')
        ip_address = request.remote_addr
        
        conn = get_db_connection()
        conn.execute("INSERT INTO audit_logs (username, ip_address, action, details) VALUES (?, ?, ?, ?)",
                     (username, ip_address, action, json.dumps(details) if details else None))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Error logging activity: {e}")

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def api_login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({"message": "Unauthorized"}), 401
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session or session.get('role') != 'admin':
            return "Unauthorized. Admin access required.", 403
        return f(*args, **kwargs)
    return decorated_function

SETTINGS_FILE = "settings.json"

class SessionContext:
    def __init__(self, session_id):
        self.session_id = session_id
        self.logs = []
        self.frame_b64 = ""
        self.stop_event = None
        self.pause_event = None
        self.active_drivers = [] # List of all drivers in this session
        self.driver_lock = threading.Lock()
        self.log_file = None
        self.running_thread = None

    def add_driver(self, driver):
        with self.driver_lock:
            if driver not in self.active_drivers:
                self.active_drivers.append(driver)

    def remove_driver(self, driver):
        with self.driver_lock:
            if driver in self.active_drivers:
                self.active_drivers.remove(driver)

    def cleanup_drivers(self):
        with self.driver_lock:
            for driver in self.active_drivers[:]:
                try:
                    driver.quit()
                except Exception:
                    pass
            self.active_drivers.clear()

# TTL cache: max 500 concurrent sessions, auto-expire after 1 hour (prevents memory DoS)
sessions = TTLCache(maxsize=500, ttl=3600)

def get_session(req):
    sid = req.args.get('session_id')
    if not sid and req.is_json:
        sid = req.json.get('session_id')
    if not sid:
        sid = 'default'
        
    if sid not in sessions:
        sessions[sid] = SessionContext(sid)
    return sessions[sid]

def get_user_settings_file(username):
    os.makedirs("data", exist_ok=True)
    return os.path.join("data", f"settings_{username}.json")

def load_settings(username):
    filename = get_user_settings_file(username)
    if not os.path.exists(filename):
        return {"spreadsheet_id": "", "sheet_name": "", "service_json_path": ""}
    with open(filename, "r") as f:
        try:
            return json.load(f)
        except json.JSONDecodeError:
            return {"spreadsheet_id": "", "sheet_name": "", "service_json_path": ""}

def save_settings(username, data):
    filename = get_user_settings_file(username)
    with open(filename, "w") as f:
        json.dump(data, f, indent=4)


@app.route("/login", methods=["GET", "POST"])
@limiter.limit("5 per minute", methods=["POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        conn = get_db_connection()
        user = conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
        conn.close()
        
        if user and check_password_hash(user['password_hash'], password):
            if user['mfa_enabled']:
                # Pre-auth state: stash user info and redirect to 2FA
                session.clear()  # Session fixation fix
                session['mfa_pre_auth_user_id'] = user['id']
                session['mfa_pre_auth_username'] = user['username']
                return render_template("login_2fa.html")
            
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            log_activity("LOGIN_SUCCESS", {"method": "password"})

            # Security: Detect default admin password to nudge for change
            if user['username'] == 'admin' and password == 'admin123':
                session['password_warning'] = True
            else:
                session.pop('password_warning', None)

            if not user['mfa_enabled']:
                return redirect(url_for('mfa_nudge'))
            return redirect(url_for('index'))
        
        log_activity("LOGIN_FAILED", {"username": username}, username=username)
        return render_template("login.html", error="Invalid credentials")
    return render_template("login.html")

@app.route("/login/2fa", methods=["POST"])
@limiter.limit("5 per minute")
def login_2fa():
    user_id = session.get('mfa_pre_auth_user_id')
    username = session.get('mfa_pre_auth_username')
    if not user_id or not username:
        return redirect(url_for('login'))
        
    code = request.form.get("code")
    conn = get_db_connection()
    user = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    conn.close()
    
    if user and user['mfa_secret']:
        totp = pyotp.TOTP(user['mfa_secret'])
        if totp.verify(code):
            # Finalize auth — regenerate session to prevent fixation
            session.clear()
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            log_activity("LOGIN_SUCCESS_MFA")
            return redirect(url_for('index'))
            
    log_activity("LOGIN_FAILED_MFA", {"username": username}, username=username)
    return render_template("login_2fa.html", error="Invalid 2FA code")

@app.route("/mfa_nudge")
def mfa_nudge():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template("mfa_nudge.html")

@app.route("/mfa_setup")
def mfa_setup_page():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template("mfa_setup.html")

@app.route("/logout")
def logout():
    log_activity("LOGOUT")
    session.clear()
    return redirect(url_for("login"))

@app.route("/api/user/settings", methods=["GET", "POST"])
@api_login_required
def user_settings():
    username = session.get('username')
    conn = get_db_connection()
    if request.method == "POST":
        data = request.json
        settings_json = json.dumps(data)
        conn.execute("UPDATE users SET settings = ? WHERE username = ?", (settings_json, username))
        conn.commit()
        conn.close()
        log_activity("SETTINGS_UPDATE", data)
        return jsonify({"message": "Settings saved successfully", "settings": data})
    
    user = conn.execute("SELECT settings FROM users WHERE username = ?", (username,)).fetchone()
    conn.close()
    settings = json.loads(user['settings']) if user and user['settings'] else {}
    return jsonify(settings)

@app.route("/api/history/<filename>")
def api_history_file(filename):
    username = session.get('username')
    if not username:
        return jsonify({"message": "Unauthorized"}), 401
    
    log_dir = os.path.abspath("logs")
    file_path = os.path.abspath(os.path.join(log_dir, filename))
    
    # ── Path traversal fix: ensure resolved path stays inside logs/ ──
    if not file_path.startswith(log_dir + os.sep):
        return jsonify({"message": "Access denied"}), 403

    if not os.path.exists(file_path) or not os.path.isfile(file_path):
        return jsonify({"message": "File not found"}), 404
    
    # Ownership check: user can only access their own logs; admin can access all
    base = os.path.basename(file_path)
    if not base.startswith(f"run_{username}_") and not base.startswith(f"archive_{username}_"):
        if session.get('role') != 'admin':
            return jsonify({"message": "Access denied to this file"}), 403
            
    return send_file(file_path, as_attachment=True)

@app.route("/login/google")
def login_google():
    redirect_uri = url_for('authorize_google', _external=True)
    return google.authorize_redirect(redirect_uri)

@app.route("/login/google/authorized")
def authorize_google():
    token = google.authorize_access_token()
    user_info = token.get('userinfo')
    if not user_info:
        return "SSO Error: No user info returned", 400
    
    email = user_info.get('email')
    
    conn = get_db_connection()

    # --- Handle pending invite token ---
    invite_token = session.pop("pending_invite_token", None)
    if invite_token:
        invite = conn.execute(
            "SELECT * FROM invite_tokens WHERE token = ? AND used = 0", (invite_token,)
        ).fetchone()
        if invite and datetime.utcnow() <= datetime.strptime(invite["expires_at"], "%Y-%m-%d %H:%M:%S"):
            # Check if user already exists
            existing = conn.execute("SELECT * FROM users WHERE username = ?", (email,)).fetchone()
            if not existing:
                conn.execute(
                    "INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)",
                    (email, generate_password_hash(secrets.token_hex(16)), invite["role"])
                )
            conn.execute("UPDATE invite_tokens SET used = 1 WHERE token = ?", (invite_token,))
            conn.commit()
            user = conn.execute("SELECT * FROM users WHERE username = ?", (email,)).fetchone()
            session.clear()  # Session fixation fix
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            conn.close()
            log_activity("LOGIN_VIA_INVITE", {"email": email, "role": invite["role"]})
            return redirect(url_for('mfa_nudge'))

    user = conn.execute("SELECT * FROM users WHERE username = ?", (email,)).fetchone()
    conn.close()
    
    if user:
        if user['mfa_enabled']:
            session.clear()  # Session fixation fix
            session['mfa_pre_auth_user_id'] = user['id']
            session['mfa_pre_auth_username'] = user['username']
            log_activity("LOGIN_SSO_PENDING_MFA", {"username": email})
            return render_template("login_2fa.html")
            
        session.clear()  # Session fixation fix
        session['user_id'] = user['id']
        session['username'] = user['username']
        session['role'] = user['role']
        log_activity("LOGIN_SUCCESS", {"method": "sso"})
        
        if not user['mfa_enabled']:
            return redirect(url_for('mfa_nudge'))
        return redirect(url_for('index'))
    else:
        return render_template("login.html", error=f"Access Denied: The Google account '{email}' is not authorized. Please ask an Administrator to whitelist you.")

@app.route("/change_password", methods=["GET", "POST"])
@login_required
def change_password():
    if request.method == "POST":
        current_password = request.form.get("current_password")
        new_password = request.form.get("new_password", "")

        if len(new_password) < 8:
            return render_template("change_password.html", error="New password must be at least 8 characters.")

        conn = get_db_connection()
        user = conn.execute("SELECT * FROM users WHERE id = ?", (session['user_id'],)).fetchone()
        
        if user and check_password_hash(user['password_hash'], current_password):
            conn.execute("UPDATE users SET password_hash = ? WHERE id = ?", (generate_password_hash(new_password), session['user_id']))
            conn.commit()
            conn.close()
            return redirect(url_for('index'))
        else:
            conn.close()
            return render_template("change_password.html", error="Incorrect current password")
            
    return render_template("change_password.html")

@app.route("/admin")
@admin_required
def admin_dashboard():
    conn = get_db_connection()
    users = conn.execute("SELECT id, username, role FROM users").fetchall()
    conn.close()
    return render_template("admin.html", users=users)

@app.route("/admin/add_user", methods=["POST"])
@admin_required
def add_user():
    username = request.form.get("username")
    password = request.form.get("password")
    role = request.form.get("role", "user")

    # ── Security: Whitelist allowed roles to prevent privilege escalation ──
    ALLOWED_ROLES = {"user", "admin"}
    if role not in ALLOWED_ROLES:
        return "Invalid role. Allowed values: user, admin.", 400

    if not username or not password:
        return "Username and password required", 400

    if len(password) < 8:
        return "Password must be at least 8 characters.", 400

    try:
        conn = get_db_connection()
        conn.execute("INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)",
                     (username, generate_password_hash(password), role))
        conn.commit()
        conn.close()
        log_activity("USER_ADDED", {"username": username, "role": role})
        return redirect(url_for('admin_dashboard'))
    except sqlite3.IntegrityError:
        return "Username already exists.", 400

@app.route("/admin/delete_user/<int:user_id>", methods=["POST"])
@admin_required
def delete_user(user_id):
    if session.get('user_id') == user_id:
        return "Cannot delete yourself.", 400
        
    conn = get_db_connection()
    target_user = conn.execute("SELECT username FROM users WHERE id = ?", (user_id,)).fetchone()
    if target_user and target_user['username'] == 'admin':
        conn.close()
        return "Cannot delete the main admin account.", 403
        
    conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()
    log_activity("USER_DELETED", {"target_username": target_user['username']} if target_user else {"target_user_id": user_id})
    return redirect(url_for('admin_dashboard'))

# ── Admin: Reset 2FA for a user ─────────────────────────────────────
@app.route("/admin/reset_mfa/<int:user_id>", methods=["POST"])
@admin_required
def admin_reset_mfa(user_id):
    conn = get_db_connection()
    target = conn.execute("SELECT username FROM users WHERE id = ?", (user_id,)).fetchone()
    if not target:
        conn.close()
        return jsonify({"error": "User not found"}), 404
    conn.execute("UPDATE users SET mfa_secret = NULL, mfa_enabled = 0 WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()
    log_activity("admin_reset_mfa", {"target_user": target["username"]})
    return jsonify({"success": True, "message": f"2FA reset for {target['username']}. They will be prompted to re-enroll on next login."})

# ── Admin: Send email invite ─────────────────────────────────────────
@app.route("/admin/invite_user", methods=["POST"])
@admin_required
def invite_user():
    data = request.get_json()
    email = (data.get("email") or "").strip()
    role = data.get("role", "user")

    if not email:
        return jsonify({"error": "Email is required"}), 400

    token = secrets.token_urlsafe(32)
    expires_at = (datetime.utcnow() + timedelta(hours=48)).strftime("%Y-%m-%d %H:%M:%S")

    conn = get_db_connection()
    conn.execute(
        "INSERT INTO invite_tokens (token, email, role, expires_at) VALUES (?, ?, ?, ?)",
        (token, email, role, expires_at)
    )
    conn.commit()
    conn.close()

    # Use the actual network IP/hostname the admin is currently accessing the dashboard from
    base_url = request.host_url.rstrip('/')
    invite_link = f"{base_url}/invite/{token}"

    html = f"""
    <div style="font-family:sans-serif; max-width:500px; margin:auto; padding:32px; background:#0f0f0f; color:#f0f0f0; border-radius:12px;">
        <h2 style="color:#6366f1;">You've been invited to Takeout Automator</h2>
        <p>An admin has invited you to join the Takeout Automator dashboard as a <strong>{role}</strong>.</p>
        <p>Click the button below to accept the invitation and sign in with your Google account:</p>
        <a href="{invite_link}" style="display:inline-block; margin:16px 0; padding:12px 24px; background:#6366f1; color:white; text-decoration:none; border-radius:8px; font-weight:600;">
            Accept Invitation &amp; Sign In
        </a>
        <p style="font-size:0.85rem; color:#888;">This link expires in 48 hours. If you didn't expect this, you can ignore this email.</p>
    </div>
    """

    print(f"\n[!] INVITATION LINK GENERATED: {invite_link}\n")

    sent = send_email(email, "You've been invited to Takeout Automator", html)
    if sent:
        log_activity("invite_sent", {"email": email, "role": role})
        return jsonify({"success": True, "message": f"Invitation sent to {email}"})
    else:
        log_activity("INVITE_EMAIL_FAILED", {"email": email, "fallback_link": invite_link})
        # Even if email fails, let's just return success for testing/local setups so admin can still use the link
        return jsonify({"success": True, "message": f"Saved locally! Check terminal or audit logs for the invite link."})

# ── Accept invite → SSO ───────────────────────────────────────────────
@app.route("/invite/<token>")
def accept_invite(token):
    conn = get_db_connection()
    invite = conn.execute(
        "SELECT * FROM invite_tokens WHERE token = ? AND used = 0", (token,)
    ).fetchone()
    conn.close()

    if not invite:
        return render_template("login.html", error="Invalid or already used invitation link.")

    if datetime.utcnow() > datetime.strptime(invite["expires_at"], "%Y-%m-%d %H:%M:%S"):
        return render_template("login.html", error="This invitation link has expired. Please ask the admin to send a new one.")

    # Store token in session so callback can validate and create user
    session["pending_invite_token"] = token
    return redirect(url_for("login_google"))

# ── API: List available drives (for multi-drive download path) ────────
@app.route("/api/drives")
@login_required
def get_drives():
    drives = []
    for letter in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
        path = f"{letter}:\\"
        if os.path.exists(path):
            try:
                import subprocess
                result = subprocess.run(
                    ['wmic', 'logicaldisk', f'where DeviceID="{letter}:"', 'get', 'Size,FreeSpace', '/format:csv'],
                    capture_output=True, text=True, timeout=3
                )
                parts = [line for line in result.stdout.strip().split('\n') if line.strip() and ',' in line]
                if parts:
                    row_parts = parts[-1].split(',')
                    free = row_parts[1].strip() if len(row_parts) > 1 else ''
                    total = row_parts[2].strip() if len(row_parts) > 2 else ''
                    drives.append({
                        "letter": letter,
                        "path": path,
                        "free_gb": round(int(free) / (1024**3), 1) if free.isdigit() else None,
                        "total_gb": round(int(total) / (1024**3), 1) if total.isdigit() else None
                    })
                else:
                    drives.append({"letter": letter, "path": path, "free_gb": None, "total_gb": None})
            except Exception:
                drives.append({"letter": letter, "path": path, "free_gb": None, "total_gb": None})
    return jsonify(drives)

@app.route("/admin/toggle_role/<int:user_id>", methods=["POST"])
@admin_required
def toggle_role(user_id):
    if session.get('user_id') == user_id:
        return "Cannot change your own role.", 400
    
    conn = get_db_connection()
    user = conn.execute("SELECT username, role FROM users WHERE id = ?", (user_id,)).fetchone()
    
    if user:
        if user['username'] == 'admin':
            conn.close()
            return "Cannot change the role of the main admin account.", 403
            
        new_role = 'admin' if user['role'] == 'user' else 'user'
        conn.execute("UPDATE users SET role = ? WHERE id = ?", (new_role, user_id))
        conn.commit()
        log_activity("USER_ROLE_TOGGLED", {"target_username": user['username'], "new_role": new_role})
        
    conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route("/admin/reset_password/<int:user_id>", methods=["POST"])
@admin_required
def admin_reset_password(user_id):
    conn = get_db_connection()
    target_user = conn.execute("SELECT username FROM users WHERE id = ?", (user_id,)).fetchone()
    
    if target_user and target_user['username'] == 'admin' and session.get('username') != 'admin':
        conn.close()
        return "Only the main admin can reset their own password.", 403
        
    new_password = request.form.get("new_password")
    if not new_password or len(new_password) < 8:
        conn.close()
        return "Password must be at least 8 characters.", 400
    conn.execute("UPDATE users SET password_hash = ? WHERE id = ?", (generate_password_hash(new_password), user_id))
    conn.commit()
    conn.close()
    log_activity("USER_PASSWORD_RESET", {"target_username": target_user['username']} if target_user else {"target_user_id": user_id})
    return redirect(url_for('admin_dashboard'))

@app.route("/admin/user_logs/<username>")
@admin_required
def admin_user_logs(username):
    if not os.path.exists("logs"):
        user_logs = []
    else:
        files = sorted(os.listdir("logs"), reverse=True)
        user_logs = [f for f in files if f.startswith(f"run_{username}_")]
    return render_template("admin_logs.html", target_user=username, logs=user_logs)

@app.route("/")
@login_required
def index():
    return render_template("index.html")

@app.route("/api/upload_csv", methods=["POST"])
@csrf.exempt
def upload_csv():
    """Import target accounts from a CSV file. Overrides existing data for the user."""
    # --- Dual Authentication: Web Session OR Agent Token ---
    username = None
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        token = auth[7:]
        conn = get_db_connection()
        agent = conn.execute("SELECT username FROM agent_registrations WHERE token = ?", (token,)).fetchone()
        conn.close()
        if not agent:
            return jsonify({"error": "Invalid agent token"}), 403
        username = agent["username"]
    else:
        if 'user_id' not in session:
            return jsonify({"error": "Unauthorized"}), 401
        username = session.get('username', 'admin')

    if 'file' not in request.files:
        return jsonify({"message": "No file part"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"message": "No selected file"}), 400
    if file and file.filename.endswith('.csv'):
        # Use utf-8-sig to handle potential Byte Order Marks (BOM) from Excel exports
        content = file.stream.read().decode("utf-8-sig", errors="ignore")
        stream = io.StringIO(content, newline=None)
        csv_input = csv.DictReader(stream)
        
        # Normalize field names to lowercase and strip whitespace for robust matching
        if csv_input.fieldnames:
            csv_input.fieldnames = [fn.strip().lower() for fn in csv_input.fieldnames]
        
        conn = get_db_connection()
        
        # --- ARCHIVE EXISTING DATA FOR THIS USER FIRST ---
        existing = conn.execute("SELECT email, password, backup_codes, create_status, download_status FROM target_accounts WHERE owner = ?", (username,)).fetchall()
        if existing:
            os.makedirs("logs", exist_ok=True)
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            archive_path = os.path.join("logs", f"archive_{username}_{timestamp}.csv")
            with open(archive_path, "w", encoding="utf-8", newline="") as f:
                cw = csv.writer(f)
                cw.writerow(['Email', 'Password', 'BackupCodes', 'CreateStatus', 'DownloadStatus'])
                for row in existing:
                    # NOTE: passwords intentionally omitted from archive for security
                    cw.writerow([row['email'], '[REDACTED]', row['backup_codes'], row['create_status'], row['download_status']])
            print(f"Archived existing database to {archive_path}")
            log_activity("CSV_ARCHIVE_AUTO", {"file": os.path.basename(archive_path)}, username=username)

        # --- OVERRIDE THIS USER'S DATA ONLY ---
        conn.execute("DELETE FROM target_accounts WHERE owner = ?", (username,))
        
        count = 0
        for row in csv_input:
            # Robust mapping for email, password, backup codes, and statuses
            email = (row.get('email') or row.get('username') or row.get('e-mail') or '').strip()
            password = (row.get('password') or row.get('pwd') or row.get('pass') or '').strip()
            backup_codes = (row.get('backupcodes') or row.get('backup codes') or row.get('backup_codes') or 
                            row.get('backup') or row.get('2fa') or row.get('mfa') or '').strip()
            
            # Extract statuses if present, default to 'Pending'
            c_status = (row.get('createstatus') or row.get('create_status') or row.get('create status') or 'Pending').strip()
            d_status = (row.get('downloadstatus') or row.get('download_status') or row.get('download status') or 'Pending').strip()
            
            if email and password:
                conn.execute("""
                    INSERT INTO target_accounts (email, password, backup_codes, create_status, download_status, owner) 
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (email, encrypt_password(password), backup_codes, c_status, d_status, username))
                count += 1
        
        conn.commit()
        conn.close()
        log_activity("CSV_UPLOAD", {"count": count}, username=username)
        return jsonify({
            "success": True, 
            "inserted": count,
            "message": f"Database overridden. Successfully imported {count} accounts!"
        })
    return jsonify({"success": False, "message": "Invalid file format. Please upload a .csv file."}), 400

@app.route("/api/add_target_account", methods=["POST"])
@csrf.exempt
def add_target_account():
    # --- Dual Authentication: Web Session OR Agent Token ---
    username = None
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        token = auth[7:]
        conn = get_db_connection()
        agent = conn.execute("SELECT username FROM agent_registrations WHERE token = ?", (token,)).fetchone()
        conn.close()
        if not agent:
            return jsonify({"error": "Invalid agent token"}), 403
        username = agent["username"]
    else:
        if 'user_id' not in session:
            return jsonify({"error": "Unauthorized"}), 401
        username = session.get('username')

    data = request.json
    email = data.get('email', '').strip()
    password = data.get('password', '').strip()
    backup_codes = data.get('backup_codes', '').strip()
    
    if not email or not password:
        return jsonify({"message": "Email and Password are required"}), 400
        
    conn = get_db_connection()
    
    # Check for duplicate
    existing = conn.execute("SELECT id FROM target_accounts WHERE email = ? AND owner = ?", (email, username)).fetchone()
    if existing:
        conn.close()
        return jsonify({"message": "Account already exists"}), 400
        
    try:
        conn.execute("INSERT INTO target_accounts (email, password, backup_codes, owner) VALUES (?, ?, ?, ?)",
                     (email, encrypt_password(password), backup_codes, username))
        conn.commit()
        log_activity("ACCOUNT_ADD_MANUAL", {"email": email, "owner": username}, username=username)
        return jsonify({"success": True, "message": "Account added successfully!"})
    except Exception as e:
        return jsonify({"success": False, "message": f"Error adding account: {str(e)}"}), 500
    finally:
        conn.close()

@app.route("/api/target_accounts", methods=["GET"])
@api_login_required
def get_target_accounts():
    source = request.args.get("source", "csv")
    username = session.get('username')
    
    if source == "sheet":
        accounts = fetch_sheet_accounts(username)
        return jsonify(accounts)
    else:
        conn = get_db_connection()
        accounts = conn.execute("SELECT id, email, backup_codes, create_status, download_status FROM target_accounts WHERE owner = ?", (username,)).fetchall()
        conn.close()
        return jsonify([dict(ix) for ix in accounts])

@app.route("/api/delete_account", methods=["POST"])
@api_login_required
def delete_target_account():
    username = session.get('username')
    payload = request.json or {}
    email = payload.get("email")
    
    if not email:
        return jsonify({"message": "Email is required"}), 400
        
    conn = get_db_connection()
    res = conn.execute("DELETE FROM target_accounts WHERE email = ? AND owner = ?", (email, username))
    conn.commit()
    conn.close()
    
    if res.rowcount > 0:
        log_activity("ACCOUNT_DELETED", {"email": email})
        return jsonify({"message": f"Account {email} deleted successfully"})
    else:
        return jsonify({"message": "Account not found or access denied"}), 404

@app.route("/api/delete_accounts_bulk", methods=["POST"])
@csrf.exempt
def delete_accounts_bulk():
    # --- Dual Authentication ---
    username = None
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        token = auth[7:]
        conn = get_db_connection()
        agent = conn.execute("SELECT username FROM agent_registrations WHERE token = ?", (token,)).fetchone()
        conn.close()
        if not agent: return jsonify({"error": "Invalid token"}), 403
        username = agent["username"]
    else:
        if 'user_id' not in session: return jsonify({"error": "Unauthorized"}), 401
        username = session.get('username')

    payload = request.json or {}
    emails = payload.get("emails", [])
    if not isinstance(emails, list):
        return jsonify({"message": "Invalid payload"}), 400
    emails = emails[:500]
    if not emails:
        return jsonify({"message": "No emails provided"}), 400
        
    conn = get_db_connection()
    placeholders = ",".join(["?" for _ in emails])
    query = f"DELETE FROM target_accounts WHERE owner = ? AND email IN ({placeholders})"
    res = conn.execute(query, [username] + emails)
    conn.commit()
    conn.close()
    
    log_activity("ACCOUNTS_DELETED_BULK", {"count": res.rowcount, "owner": username})
    return jsonify({"success": True, "deleted": res.rowcount})

@app.route("/api/settings", methods=["GET"])
@api_login_required
def get_settings():
    username = session.get('username')
    settings = load_settings(username)
    
    # Check MFA status from DB
    conn = get_db_connection()
    user = conn.execute("SELECT mfa_enabled FROM users WHERE username = ?", (username,)).fetchone()
    conn.close()
    
    settings['_mfa_enabled'] = bool(user['mfa_enabled']) if user else False
    return jsonify(settings)

@app.route("/api/save_settings", methods=["POST"])
@api_login_required
def update_settings():
    username = session.get('username')
    
    current_settings = load_settings(username)
    current_settings["spreadsheet_id"] = request.form.get("spreadsheet_id", "")
    current_settings["sheet_name"] = request.form.get("sheet_name", "")
    
    if 'credentials_file' in request.files:
        creds_file = request.files['credentials_file']
        if creds_file.filename != '':
            if not creds_file.filename.endswith('.json'):
                return jsonify({"message": "Credentials file must be a .json file."}), 400
            creds_path = os.path.join("data", f"credentials_{username}.json")
            raw = creds_file.read()
            try:
                json.loads(raw)  # Validate it's real JSON before saving
            except ValueError:
                return jsonify({"message": "Credentials file is not valid JSON."}), 400
            with open(creds_path, 'wb') as f:
                f.write(raw)
            current_settings["service_json_path"] = creds_path
        
    save_settings(username, current_settings)
    log_activity("SETTINGS_UPDATED", {"username": username})
    return jsonify({"message": "Settings saved", "settings": current_settings})


@app.route("/api/start", methods=["POST"])
@api_login_required
def start_automation():
    session_ctx = get_session(request)
    if session_ctx.running_thread and session_ctx.running_thread.is_alive():
        return jsonify({"message": "Already running"}), 400
    
    session_ctx.logs.clear()
    session_ctx.frame_b64 = ""

    payload = request.json or {}
    mode = payload.get("mode", "create")
    data_source = payload.get("data_source", "csv")
    browser_type = payload.get("browser_type", "chrome")
    
    username = session.get('username', 'unknown')
    # Load global/file-based settings
    run_settings = load_settings(username)
    
    # Load user-specific settings from DB and merge
    conn = get_db_connection()
    user_row = conn.execute("SELECT settings FROM users WHERE username = ?", (username,)).fetchone()
    conn.close()
    if user_row and user_row['settings']:
        db_settings = json.loads(user_row['settings'])
        run_settings.update(db_settings)
    
    run_settings['browser_type'] = browser_type
    # Pass showBrowser from frontend — defaults to True (visible) if not provided
    run_settings['showBrowser'] = payload.get('showBrowser', True)
    
    if data_source == "sheet":
        if not run_settings.get("spreadsheet_id"):
            return jsonify({"message": "Spreadsheet ID is missing in Settings. Please provide it."}), 400
        if not run_settings.get("service_json_path") or not os.path.exists(run_settings.get("service_json_path")):
            return jsonify({"message": "Service Account JSON is missing or corrupted. Please upload your credentials file in Settings."}), 400

    target_agent = payload.get("target_agent", "local")
    
    if target_agent != "local":
        # Enqueue for remote agent
        conn = get_db_connection()
        agent = conn.execute("SELECT * FROM agent_registrations WHERE agent_id = ?", (target_agent,)).fetchone()
        if not agent:
            conn.close()
            return jsonify({"message": f"Agent {target_agent} not found"}), 404
        
        # Get all accounts for this user
        accounts = conn.execute("SELECT email FROM target_accounts WHERE owner = ?", (username,)).fetchall()
        if not accounts:
            conn.close()
            return jsonify({"message": "No accounts found for automation"}), 400
        
        for acc in accounts:
            conn.execute(
                "INSERT INTO agent_jobs (agent_id, email, mode, status, settings) VALUES (?, ?, ?, 'pending', ?)",
                (target_agent, acc["email"], mode, json.dumps(run_settings))
            )
        conn.commit()
        conn.close()
        
        log_activity("REMOTE_AUTOMATION_QUEUED", {"agent_id": target_agent, "mode": mode, "count": len(accounts)})
        return jsonify({"message": f"Successfully queued {len(accounts)} tasks for Agent {agent['username']} ({agent['hostname']})"})

    # Local execution (existing logic)
    session_ctx.stop_event = threading.Event()
    session_ctx.pause_event = threading.Event()
    session_ctx.pause_event.set()
    
    os.makedirs("logs", exist_ok=True)
    clean_sid = "".join(c for c in session_ctx.session_id if c.isalnum() or c in "_-")
    session_ctx.log_file = os.path.join("logs", time.strftime(f"run_{username}_{clean_sid}_%Y%m%d_%H%M%S.log"))
    
    if mode == "download":
        target_fn = takeout_downloader.run_all
    else:
        target_fn = takeout_runner.run_all

    run_settings['_run_owner'] = username
    
    session_ctx.running_thread = threading.Thread(target=target_fn, args=(run_settings, session_ctx, data_source))
    session_ctx.running_thread.start()
    log_activity("LOCAL_AUTOMATION_START", {"mode": mode, "source": data_source})
    return jsonify({"message": f"Local automation started in {mode} mode using {data_source} data source"})

@app.route("/api/status", methods=["GET"])
@api_login_required
def get_run_status():
    session_ctx = get_session(request)
    is_running = False
    if session_ctx.running_thread and session_ctx.running_thread.is_alive():
        is_running = True
    
    # Also check if any drivers are still hanging around
    has_active_drivers = False
    with session_ctx.driver_lock:
        if session_ctx.active_drivers:
            has_active_drivers = True

    return jsonify({
        "is_running": is_running or has_active_drivers,
        "is_thread_alive": is_running,
        "active_drivers_count": len(session_ctx.active_drivers) if hasattr(session_ctx, 'active_drivers') else 0,
        "session_id": session_ctx.session_id
    })

@app.route("/api/stop", methods=["POST"])
@api_login_required
def stop_automation():
    session_ctx = get_session(request)
    if session_ctx.stop_event:
        session_ctx.stop_event.set()
        if session_ctx.pause_event:
            session_ctx.pause_event.set()
        
        # Robust cleanup of ALL drivers tracked in this session
        session_ctx.cleanup_drivers()
        
        log_activity("AUTOMATION_STOP")
        return jsonify({"message": "Stopping all automation sessions..."})
    return jsonify({"message": "Not running"})

@app.route("/api/reset", methods=["POST"])
@api_login_required
def reset_session():
    """Emergency reset: clears the session state and kills all drivers."""
    session_ctx = get_session(request)
    if session_ctx.stop_event:
        session_ctx.stop_event.set()
    
    session_ctx.cleanup_drivers()
    session_ctx.logs.append("[SYSTEM] Emergency reset triggered. Cleared all active drivers.")
    log_activity("SESSION_RESET")
    return jsonify({"message": "Session reset successfully. All drivers terminated."})

@app.route("/api/pause", methods=["POST"])
@api_login_required
def pause_automation():
    session_ctx = get_session(request)
    if session_ctx.pause_event:
        session_ctx.pause_event.clear()
        log_activity("AUTOMATION_PAUSE")
        return jsonify({"message": "Automation PAUSED. (Currently running step will finish before pausing)"})
    return jsonify({"message": "Not running"})

@app.route("/api/resume", methods=["POST"])
@api_login_required
def resume_automation():
    session_ctx = get_session(request)
    if session_ctx.pause_event:
        session_ctx.pause_event.set()
        log_activity("AUTOMATION_RESUME")
        return jsonify({"message": "Automation RESUMED."})
    return jsonify({"message": "Not running"})

@app.route("/api/user/mfa/setup", methods=["POST"])
@api_login_required
def mfa_setup():
    username = session.get('username')
    secret = pyotp.random_base32()
    # Stash secret in session until verified
    session['mfa_pending_secret'] = secret
    
    # Generate TOTP URI for QR code
    totp = pyotp.TOTP(secret)
    provision_uri = totp.provisioning_uri(name=username, issuer_name="TakeoutDashboard")
    
    return jsonify({
        "secret": secret,
        "provision_uri": provision_uri
    })

@app.route("/api/user/mfa/verify", methods=["POST"])
@api_login_required
@limiter.limit("5 per minute")  # Prevent TOTP brute-force attempts
def mfa_verify():
    username = session.get('username')
    pending_secret = session.get('mfa_pending_secret')
    if not pending_secret:
        return jsonify({"message": "No pending 2FA setup"}), 400
        
    code = request.json.get("code")
    totp = pyotp.TOTP(pending_secret)
    if totp.verify(code):
        conn = get_db_connection()
        conn.execute("UPDATE users SET mfa_secret = ?, mfa_enabled = 1 WHERE username = ?", (pending_secret, username))
        conn.commit()
        conn.close()
        session.pop('mfa_pending_secret')
        log_activity("MFA_ENABLED")
        return jsonify({"message": "2FA successfully enabled!"})
    
    return jsonify({"message": "Invalid verification code"}), 400

@app.route("/api/user/mfa/disable", methods=["POST"])
@api_login_required
def mfa_disable():
    username = session.get('username')
    # Require password confirmation to disable MFA
    password = (request.json or {}).get('password', '')
    conn = get_db_connection()
    user = conn.execute("SELECT password_hash FROM users WHERE username = ?", (username,)).fetchone()
    if not user or not check_password_hash(user['password_hash'], password):
        conn.close()
        return jsonify({"message": "Incorrect password. Please confirm your current password to disable 2FA."}), 403
    conn.execute("UPDATE users SET mfa_enabled = 0 WHERE username = ?", (username,))
    conn.commit()
    conn.close()
    log_activity("MFA_DISABLED")
    return jsonify({"message": "2FA disabled."})
@app.route("/api/show_browser", methods=["POST"])
@api_login_required
def show_browser():
    session_ctx = get_session(request)
    if hasattr(session_ctx, 'driver') and session_ctx.driver:
        try:
            session_ctx.driver.set_window_position(0, 0)
            session_ctx.driver.maximize_window()
            return jsonify({"message": "Browser teleported to your screen!"})
        except Exception as e:
            err_str = str(e).lower()
            # If the driver session is dead (connection refused, no such window, etc.),
            # clear the reference so we don't keep spamming connection errors.
            if any(k in err_str for k in ['connection refused', 'failed to establish', 'no such window', 'no such session', 'target window already closed', 'unable to connect']):
                session_ctx.driver = None
                return jsonify({"message": "Browser session ended. No active browser."}), 200
            return jsonify({"message": f"Error showing browser: {e}"}), 400
    return jsonify({"message": "No active browser session right now."}), 200

@app.route("/api/hide_browser", methods=["POST"])
@api_login_required
def hide_browser():
    session_ctx = get_session(request)
    if hasattr(session_ctx, 'driver') and session_ctx.driver:
        try:
            # Try to restore window if maximized, but swallow errors as long as position works
            try:
                session_ctx.driver.set_window_size(1200, 900)
            except Exception:
                pass
                
            session_ctx.driver.set_window_position(-3000, 0)
            return jsonify({"message": "Browser banished off-screen!"})
        except Exception as e:
            err_str = str(e).lower()
            # If the driver session is dead, clear the reference silently
            if any(k in err_str for k in ['connection refused', 'failed to establish', 'no such window', 'no such session', 'target window already closed', 'unable to connect']):
                session_ctx.driver = None
                return jsonify({"message": "Browser session ended."}), 200
            # If it's already off-screen or the move worked despite the error, return success
            if "window state" in err_str:
                return jsonify({"message": "Browser position updated (swallowed state error)"})
            return jsonify({"message": f"Error hiding browser: {e}"}), 400
    return jsonify({"message": "No active browser session right now."}), 200

@app.route("/api/logs/stream")
@api_login_required
def stream_logs():
    sid = request.args.get('session_id', 'default')
    def generate(_sid):
        session_ctx = sessions.get(_sid, SessionContext(_sid))
        last_index = 0
        yield "data: --- NEW STREAM CONNECTED ---\n\n"
        while True:
            current_len = len(session_ctx.logs)
            if last_index > current_len:
                last_index = 0
                yield "data: --- NEW RUN STARTED ---\n\n"
            
            while last_index < current_len:
                line = session_ctx.logs[last_index]
                yield f"data: {line.strip()}\n\n"
                last_index += 1
            time.sleep(0.5)
    return Response(generate(sid), mimetype="text/event-stream", headers={
        "Cache-Control": "no-cache",
        "Connection": "keep-alive",
        "X-Accel-Buffering": "no"
    })

@app.route("/api/video/frame")
@api_login_required
def serve_frame():
    session_ctx = get_session(request)
    return Response(session_ctx.frame_b64, mimetype="text/plain")

@app.route("/api/export/logs", methods=["GET"])
@api_login_required
def export_logs():
    session_ctx = get_session(request)
    return Response("\n".join(session_ctx.logs), mimetype="text/plain", headers={"Content-disposition": "attachment; filename=takeout_execution.log"})

@app.route("/api/export/report", methods=["GET"])
@api_login_required
def export_report():
    username = session.get('username', 'admin')
    conn = get_db_connection()
    accounts = conn.execute("SELECT email, backup_codes, create_status, download_status FROM target_accounts WHERE owner = ?", (username,)).fetchall()
    conn.close()
    
    si = io.StringIO()
    cw = csv.writer(si)
    cw.writerow(['Email', 'BackupCodes', 'CreateStatus', 'DownloadStatus'])
    for acc in accounts:
        # Reconstruct exactly what happened
        cw.writerow([acc['email'], acc['backup_codes'], acc['create_status'], acc['download_status']])
        
    return Response(si.getvalue(), mimetype="text/csv", headers={"Content-disposition": "attachment; filename=takeout_report.csv"})

@app.route("/api/history", methods=["GET"])
@api_login_required
def get_history():
    if not os.path.exists("logs"): return jsonify([])
    files = sorted(os.listdir("logs"), reverse=True)
    # Filter to only show run logs and archive reports
    files = [f for f in files if f.startswith("run_") or f.startswith("archive_")]
    
    if session.get('role') != 'admin':
        username = session.get('username', '')
        # Non-admin users can only see their own run logs and their own archives
        files = [f for f in files if f.startswith(f"run_{username}_") or f.startswith(f"archive_{username}_")]
    return jsonify(files)

@app.route("/api/admin/audit_logs", methods=["GET"])
@admin_required
def get_audit_logs():
    conn = get_db_connection()
    logs = conn.execute("SELECT * FROM audit_logs ORDER BY id DESC LIMIT 200").fetchall()
    conn.close()
    return jsonify([dict(ix) for ix in logs])

@app.route("/api/admin/audit_logs/<username>", methods=["GET"])
@admin_required
def get_user_audit_logs(username):
    conn = get_db_connection()
    logs = conn.execute("SELECT * FROM audit_logs WHERE username = ? ORDER BY id DESC LIMIT 500", (username,)).fetchall()
    conn.close()
    return jsonify([dict(ix) for ix in logs])

@app.route("/api/admin/audit_logs/export/<fmt>", methods=["GET"])
@admin_required
def export_audit_logs(fmt):
    conn = get_db_connection()
    logs = conn.execute("SELECT * FROM audit_logs ORDER BY id DESC").fetchall()
    conn.close()

    if fmt == "csv":
        si = io.StringIO()
        cw = csv.writer(si)
        cw.writerow(['ID', 'Timestamp', 'Username', 'IP', 'Action', 'Details'])
        for log in logs:
            cw.writerow([log['id'], log['timestamp'], log['username'], log['ip_address'], log['action'], log['details']])
        return Response(si.getvalue(), mimetype="text/csv", headers={"Content-disposition": "attachment; filename=audit_logs.csv"})

    elif fmt == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Logs"
        ws.append(['ID', 'Timestamp', 'Username', 'IP', 'Action', 'Details'])
        for log in logs:
            ws.append([log['id'], log['timestamp'], log['username'], log['ip_address'], log['action'], log['details']])
        
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return Response(out.read(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-disposition": "attachment; filename=audit_logs.xlsx"})

    elif fmt == "pdf":
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 10, "Takeout Automator - Audit Logs", ln=True, align='C')
        pdf.ln(5)
        
        pdf.set_font("Arial", 'B', 8)
        # Table Header
        pdf.cell(10, 8, "ID", 1)
        pdf.cell(35, 8, "Timestamp", 1)
        pdf.cell(35, 8, "User", 1)
        pdf.cell(25, 8, "IP", 1)
        pdf.cell(30, 8, "Action", 1)
        pdf.cell(55, 8, "Details", 1)
        pdf.ln()

        pdf.set_font("Arial", '', 7)
        for log in logs:
            # Simple truncation for PDF fit
            details = (log['details'] or "")[:40]
            pdf.cell(10, 7, str(log['id']), 1)
            pdf.cell(35, 7, str(log['timestamp']), 1)
            pdf.cell(35, 7, str(log['username']), 1)
            pdf.cell(25, 7, str(log['ip_address']), 1)
            pdf.cell(30, 7, str(log['action']), 1)
            pdf.cell(55, 7, details, 1)
            pdf.ln()

        return Response(pdf.output(), mimetype="application/pdf", headers={"Content-disposition": "attachment; filename=audit_logs.pdf"})

    return "Invalid format", 400

@app.route("/api/admin/audit/backup_history_report/<fmt>", methods=["GET"])
@admin_required
def export_backup_history(fmt):
    """Compliance Report: Exports backup jobs filtered by date range."""
    start_date = request.args.get('start')
    end_date = request.args.get('end')

    # Security: Validate date formats to prevent injection or malformed queries
    try:
        if start_date:
            datetime.strptime(start_date, '%Y-%m-%d')
        if end_date:
            datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400

    conn = get_db_connection()
    if start_date and end_date:
        start = f"{start_date} 00:00:00"
        end   = f"{end_date} 23:59:59"
        jobs = conn.execute("""
            SELECT j.*, a.username as agent_user
            FROM agent_jobs j
            LEFT JOIN agent_registrations a ON j.agent_id = a.agent_id
            WHERE (j.completed_at BETWEEN ? AND ?) OR (j.assigned_at BETWEEN ? AND ?)
            ORDER BY j.id DESC
        """, (start, end, start, end)).fetchall()
        period_label = f"{start_date} to {end_date}"
    else:
        one_year_ago = (datetime.utcnow() - timedelta(days=365)).strftime("%Y-%m-%d %H:%M:%S")
        jobs = conn.execute("""
            SELECT j.*, a.username as agent_user
            FROM agent_jobs j
            LEFT JOIN agent_registrations a ON j.agent_id = a.agent_id
            WHERE j.completed_at >= ? OR j.assigned_at >= ?
            ORDER BY j.id DESC
        """, (one_year_ago, one_year_ago)).fetchall()
        period_label = "Last 12 Months"
    conn.close()

    if fmt == "csv":
        si = io.StringIO()
        cw = csv.writer(si)
        cw.writerow(['Job ID', 'Timestamp', 'Agent User', 'Target Email', 'Mode', 'Status'])
        for j in jobs:
            agent_user = j['agent_user'] if j['agent_user'] else 'server'
            cw.writerow([j['id'], j['completed_at'] or j['assigned_at'], agent_user, j['email'], j['mode'], j['status']])
        output = si.getvalue()
        return Response(
            output,
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename=backup_report_{period_label.replace(' ', '_')}.csv"}
        )

    elif fmt == "pdf":
        try:
            from fpdf import FPDF
            pdf = FPDF(orientation='L', unit='mm', format='A4')  # Landscape for more columns
            pdf.add_page()
            pdf.set_font("Arial", 'B', 14)
            pdf.cell(0, 10, "Takeout Automator - Backup Compliance Report", ln=True, align='C')
            pdf.set_font("Arial", 'I', 10)
            pdf.cell(0, 8, f"Period: {period_label}  |  Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", ln=True, align='C')
            pdf.ln(5)

            # Summary line
            pdf.set_font("Arial", 'B', 10)
            done  = sum(1 for j in jobs if j['status'] == 'done')
            failed = sum(1 for j in jobs if j['status'] == 'failed')
            pdf.cell(0, 8, f"Total Jobs: {len(jobs)}   |   Done: {done}   |   Failed: {failed}", ln=True)
            pdf.ln(3)

            # Table header
            pdf.set_fill_color(30, 41, 59)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Arial", 'B', 8)
            col_w = [15, 45, 60, 25, 30, 90]
            headers = ["ID", "Timestamp", "Target Email", "Mode", "Status", "Agent User"]
            for i, h in enumerate(headers):
                pdf.cell(col_w[i], 8, h, 1, 0, 'C', True)
            pdf.ln()

            # Table rows
            pdf.set_text_color(0, 0, 0)
            pdf.set_font("Arial", '', 7)
            fill = False
            pdf.set_fill_color(240, 242, 245)
            for j in jobs:
                ts    = (j['completed_at'] or j['assigned_at'] or "")[:19]
                email = (j['email'] or "")[:35]
                agent = (j['agent_user'] if j['agent_user'] else "server")[:40]
                status = j['status'] or ""
                # Color code status
                if status == 'done':
                    pdf.set_text_color(16, 185, 129)
                elif status == 'failed':
                    pdf.set_text_color(239, 68, 68)
                else:
                    pdf.set_text_color(100, 100, 100)

                pdf.cell(col_w[0], 6, str(j['id']), 1, 0, 'C', fill)
                pdf.set_text_color(0, 0, 0)
                pdf.cell(col_w[1], 6, ts, 1, 0, 'C', fill)
                pdf.cell(col_w[2], 6, email, 1, 0, 'L', fill)
                pdf.cell(col_w[3], 6, str(j['mode']), 1, 0, 'C', fill)
                if status == 'done':
                    pdf.set_text_color(16, 185, 129)
                elif status == 'failed':
                    pdf.set_text_color(239, 68, 68)
                pdf.cell(col_w[4], 6, status, 1, 0, 'C', fill)
                pdf.set_text_color(0, 0, 0)
                pdf.cell(col_w[5], 6, agent, 1, 0, 'L', fill)
                pdf.ln()
                fill = not fill

            pdf_bytes = bytes(pdf.output())
            response = Response(pdf_bytes, mimetype="application/pdf")
            response.headers["Content-Disposition"] = f"attachment; filename=backup_report_{period_label.replace(' ', '_')}.pdf"
            response.headers["Content-Type"] = "application/pdf"
            return response

        except ImportError:
            return jsonify({"error": "PDF generation requires the 'fpdf2' library. Run: pip install fpdf2"}), 500
        except Exception as e:
            import traceback; traceback.print_exc()
            return jsonify({"error": f"PDF generation failed: {str(e)}"}), 500

    return "Invalid format", 400

@app.route("/api/admin/active_sessions", methods=["GET"])
@admin_required
def get_active_sessions():
    # In a real app we'd track sessions in DB. 
    # For now, we return active automation sessions from our global dictionary.
    active = []
    for sid, ctx in sessions.items():
        if ctx.running_thread and ctx.running_thread.is_alive():
            active.append({
                "session_id": sid,
                "owner": getattr(ctx, 'owner', 'unknown'), # Note: we should set owner on ctx
                "start_time": getattr(ctx, 'start_time', 'N/A')
            })
    return jsonify(active)
@api_login_required
def get_history_file(filename):
    username = session.get('username', '')
    if session.get('role') != 'admin':
        # Non-admin can only access their own files
        is_own_run = filename.startswith(f"run_{username}_")
        is_own_archive = filename.startswith(f"archive_{username}_")
        if not is_own_run and not is_own_archive:
            return "Unauthorized", 403
    
    path = os.path.join("logs", filename)
    if os.path.exists(path) and ".." not in path:
        if filename.endswith(".csv"):
            return send_file(path, as_attachment=True, download_name=filename)
        with open(path, "r", encoding="utf-8") as f:
            return Response(f.read(), mimetype="text/plain")
    return "Not found", 404


# ═══════════════════════════════════════════════════════════
# AGENT API — routes called by the Windows agent on team laptops
# ═══════════════════════════════════════════════════════════

def require_agent_token(f):
    """Decorator: validates agent token from Authorization header."""
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.headers.get("Authorization", "")
        if not auth.startswith("Bearer "):
            return jsonify({"error": "Missing token"}), 401
        token = auth[7:]
        conn = get_db_connection()
        agent = conn.execute(
            "SELECT * FROM agent_registrations WHERE token = ?", (token,)
        ).fetchone()
        conn.close()
        if not agent:
            return jsonify({"error": "Invalid agent token"}), 403
        request.agent = dict(agent)
        return f(*args, **kwargs)
    return decorated


@app.route("/api/agent/register", methods=["POST"])
@csrf.exempt
def agent_register():
    """Agent first-run: exchanges invite token for an agent API token."""
    data = request.get_json()
    invite_token = data.get("invite_token", "")
    agent_id = data.get("agent_id", "")      # UUID generated on agent machine
    hostname = data.get("hostname", "unknown")

    conn = get_db_connection()
    invite = conn.execute(
        "SELECT * FROM invite_tokens WHERE token = ?", (invite_token,)
    ).fetchone()

    if not invite:
        conn.close()
        return jsonify({"error": "Invalid or expired invite token"}), 403

    # We allow the token to be reused to register multiple agents or to
    # reconnect an agent after the user clicks 'Disconnect' and loses their config

    if datetime.utcnow() > datetime.strptime(invite["expires_at"], "%Y-%m-%d %H:%M:%S"):
        conn.close()
        return jsonify({"error": "Invite token has expired"}), 403

    # Mark invite used, create/update agent registration
    agent_token = secrets.token_urlsafe(40)
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    conn.execute("UPDATE invite_tokens SET used = 1 WHERE token = ?", (invite_token,))
    conn.execute("""
        INSERT INTO agent_registrations (agent_id, username, token, hostname, last_seen)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(agent_id) DO UPDATE SET token=excluded.token, last_seen=excluded.last_seen
    """, (agent_id, invite["email"], agent_token, hostname, now))
    conn.commit()

    # Ensure user account exists for the invited email
    exists = conn.execute("SELECT id FROM users WHERE username = ?", (invite["email"],)).fetchone()
    if not exists:
        conn.execute(
            "INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)",
            (invite["email"], generate_password_hash(secrets.token_hex(16)), invite["role"])
        )
        conn.commit()

    conn.close()
    log_activity("AGENT_REGISTERED", {"agent_id": agent_id, "hostname": hostname}, username=invite["email"])
    return jsonify({"success": True, "agent_token": agent_token, "username": invite["email"]})





@app.route("/api/admin/agents/delete/<agent_id>", methods=["POST"])
@api_login_required
def admin_delete_agent(agent_id):
    """Admin can delete a registered agent."""
    conn = get_db_connection()
    conn.execute("DELETE FROM agent_registrations WHERE agent_id = ?", (agent_id,))
    conn.commit()
    conn.close()
    log_activity("AGENT_DELETE", {"agent_id": agent_id})
    return jsonify({"message": "Agent removed successfully"})


@app.route("/api/agent/heartbeat", methods=["POST"])
@csrf.exempt
@require_agent_token
def agent_heartbeat():
    """Agent POSTs every 30s to show it's online."""
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db_connection()
    conn.execute(
        "UPDATE agent_registrations SET last_seen = ? WHERE agent_id = ?",
        (now, request.agent["agent_id"])
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/agent/jobs/next", methods=["GET"])
@require_agent_token
def agent_jobs_next():
    """Returns the next pending job for this agent, or 204 if none."""
    agent_id = request.agent["agent_id"]
    conn = get_db_connection()
    job = conn.execute(
        "SELECT * FROM agent_jobs WHERE agent_id = ? AND status = 'pending' ORDER BY id LIMIT 1",
        (agent_id,)
    ).fetchone()
    if not job:
        conn.close()
        return jsonify({"job": None}), 200

    conn.execute("UPDATE agent_jobs SET status = 'running' WHERE id = ?", (job["id"],))
    conn.commit()
    conn.close()
    return jsonify({"job": dict(job)})


@app.route("/api/agent/jobs/report", methods=["POST"])
@csrf.exempt
@require_agent_token
def agent_jobs_report():
    """Agent POSTs job completion: {job_id, status, logs}."""
    data = request.get_json()
    job_id = data.get("job_id")
    status = data.get("status", "done")   # 'done' or 'failed'
    logs = data.get("logs", "")
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")

    conn = get_db_connection()
    # Ownership check: only the reporting agent can update its own jobs
    job_owner = conn.execute(
        "SELECT agent_id FROM agent_jobs WHERE id = ?", (job_id,)
    ).fetchone()
    if not job_owner or job_owner["agent_id"] != request.agent["agent_id"]:
        conn.close()
        return jsonify({"error": "Job not found or access denied"}), 403
    conn.execute(
        "UPDATE agent_jobs SET status = ?, logs = ?, completed_at = ? WHERE id = ? AND agent_id = ?",
        (status, logs, now, job_id, request.agent["agent_id"])
    )
    conn.commit()
    conn.close()
    
    log_activity("AGENT_JOB_COMPLETED", {"job_id": job_id, "status": status}, username=request.agent["username"])
    return jsonify({"ok": True})


@app.route("/api/agent/status/update", methods=["POST"])
@require_agent_token
def agent_status_update():
    """Agent: updates create_status or download_status for an account."""
    data = request.get_json()
    email = data.get("email")
    status = data.get("status")
    mode = data.get("mode", "create") # 'create' or 'download'
    username = request.agent["username"]

    conn = get_db_connection()
    # Sanitize col explicitly
    allowed_cols = {"create": "create_status", "download": "download_status"}
    col = allowed_cols.get(mode)
    if not col:
        return jsonify({"error": "Invalid mode"}), 400
        
    conn.execute(f"UPDATE target_accounts SET {col} = ? WHERE email = ? AND owner = ?", 
                 (status, email, username))
    conn.commit()
    conn.close()
    
    if 'socketio' in globals():
        socketio.emit('agent_status_update', {
            "email": email,
            "status": status,
            "mode": mode,
            "agent_id": request.agent.get("agent_id")
        })

    log_activity("AGENT_RUN", {"email": email, "status": status, "mode": mode}, username=username)
    return jsonify({"ok": True})


@app.route("/api/agent/backup_codes/update", methods=["POST"])
@require_agent_token
def agent_backup_codes_update():
    """Update backup codes for an account (e.g., consume a used code)."""
    data = request.get_json()
    email = data.get("email")
    codes = data.get("backup_codes")
    username = request.agent["username"] # Get username from the agent token

    conn = get_db_connection()
    conn.execute("UPDATE target_accounts SET backup_codes = ? WHERE email = ? AND owner = ?",
                 (codes, email, username))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/api/agent/credentials", methods=["GET"])
@require_agent_token
def agent_get_credentials():
    """Agent: returns accounts for automation."""
    username = request.agent["username"]
    conn = get_db_connection()
    rows = conn.execute("SELECT email, password, backup_codes, create_status, download_status FROM target_accounts WHERE owner = ?", (username,)).fetchall()
    conn.close()
    
    accounts = []
    for r in rows:
        accounts.append({
            "email": r["email"],
            "password": decrypt_password(r["password"]),
            "backup_codes": r["backup_codes"],
            "create_status": r["create_status"],
            "download_status": r["download_status"]
        })
    return jsonify(accounts)


@app.route("/api/agent/accounts/export")
@require_agent_token
def agent_export_accounts():
    """Export target accounts to CSV for the agent."""
    username = request.agent["username"]
    conn = get_db_connection()
    rows = conn.execute("SELECT email, backup_codes, create_status, download_status FROM target_accounts WHERE owner = ?", (username,)).fetchall()
    conn.close()
    
    si = io.StringIO()
    cw = csv.writer(si)
    cw.writerow(['Email', 'BackupCodes', 'CreateStatus', 'DownloadStatus'])
    for r in rows:
        cw.writerow([r['email'], r['backup_codes'], r['create_status'], r['download_status']])
    
    output = make_response(si.getvalue())
    output.headers["Content-Disposition"] = f"attachment; filename=target_accounts_{username}.csv"
    output.headers["Content-type"] = "text/csv"
    return output


@app.route("/api/agents/status")
@admin_required
def agents_status():
    """Admin: returns all agents with online status and latest job."""
    conn = get_db_connection()
    agents = conn.execute("SELECT * FROM agent_registrations ORDER BY last_seen DESC").fetchall()
    result = []
    cutoff = datetime.utcnow()
    for agent in agents:
        last_seen = agent["last_seen"]
        online = False
        if last_seen:
            try:
                delta = (cutoff - datetime.strptime(last_seen, "%Y-%m-%d %H:%M:%S")).total_seconds()
                online = delta < 60
            except Exception:
                pass
        latest_job = conn.execute(
            "SELECT * FROM agent_jobs WHERE agent_id = ? ORDER BY id DESC LIMIT 1",
            (agent["agent_id"],)
        ).fetchone()
        result.append({
            "agent_id": agent["agent_id"],
            "username": agent["username"],
            "hostname": agent["hostname"],
            "online": online,
            "last_seen": last_seen,
            "latest_job": dict(latest_job) if latest_job else None
        })
    conn.close()
    return jsonify(result)


@app.route("/api/agent/accounts")
@require_agent_token
def agent_get_accounts():
    """Returns the agent user's own target accounts."""
    username = request.agent["username"]
    conn = get_db_connection()
    rows = conn.execute(
        "SELECT email, backup_codes, create_status, download_status FROM target_accounts WHERE owner = ?",
        (username,)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/agent/stats")
@require_agent_token
def agent_stats():
    """Returns counts of accounts by status for the agent dashboard cards."""
    username = request.agent["username"]
    conn = get_db_connection()
    stats = conn.execute("""
        SELECT 
            SUM(CASE WHEN create_status = 'Done' OR create_status LIKE '%✅%' THEN 1 ELSE 0 END) as cs,
            SUM(CASE WHEN create_status LIKE '%❌%' OR LOWER(create_status) LIKE '%failed%' THEN 1 ELSE 0 END) as cf,
            SUM(CASE WHEN download_status = 'Done' OR download_status LIKE '%✅%' THEN 1 ELSE 0 END) as ds,
            SUM(CASE WHEN download_status LIKE '%❌%' OR LOWER(download_status) LIKE '%failed%' THEN 1 ELSE 0 END) as df
        FROM target_accounts 
        WHERE owner = ?
    """, (username,)).fetchone()
    conn.close()
    return jsonify({
        "create_success": stats["cs"] or 0,
        "create_failed": stats["cf"] or 0,
        "download_success": stats["ds"] or 0,
        "download_failed": stats["df"] or 0
    })


@app.route("/api/agent/queue_jobs", methods=["POST"])
@csrf.exempt
@require_agent_token
def agent_queue_jobs():
    """Agent queues its accounts for processing."""
    data = request.get_json()
    mode = data.get("mode", "create")     # 'create' or 'download'
    agent_id = request.agent["agent_id"]
    username = request.agent["username"]

    conn = get_db_connection()
    rows = conn.execute(
        "SELECT email FROM target_accounts WHERE owner = ?", (username,)
    ).fetchall()
    for row in rows:
        conn.execute(
            "INSERT INTO agent_jobs (agent_id, email, mode, status) VALUES (?, ?, ?, 'pending')",
            (agent_id, row["email"], mode)
        )
    conn.commit()
    conn.close()
    return jsonify({"queued": len(rows)})


@app.route("/api/user/avatar", methods=["POST"])
@api_login_required
def upload_avatar():
    if 'avatar' not in request.files:
        return jsonify({"message": "No file part"}), 400
    file = request.files['avatar']
    if file.filename == '':
        return jsonify({"message": "No selected file"}), 400

    ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
    ALLOWED_MIME_TYPES = {'png', 'jpeg', 'gif', 'webp', 'rgbe'}  # imghdr types

    if file:
        # Security: Enforce file size limit (2MB) before processing
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)
        if file_size > 2 * 1024 * 1024:
            return jsonify({"message": "Avatar file too large (max 2MB)."}), 400

        username = session.get('username')
        ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
        if ext not in ALLOWED_EXTENSIONS:
            return jsonify({"message": "Invalid file extension. Only PNG, JPG, GIF, WEBP allowed."}), 400

        # Read and validate magic bytes
        file_bytes = file.read()
        import imghdr
        detected = imghdr.what(None, h=file_bytes)
        if detected not in ALLOWED_MIME_TYPES and detected != 'jpeg':
            return jsonify({"message": "File content does not match an image type."}), 400

        filename = f"avatar_{username}_{int(time.time())}.{ext}"
        os.makedirs(os.path.join("static", "avatars"), exist_ok=True)
        filepath = os.path.join("static", "avatars", filename)
        with open(filepath, 'wb') as f:
            f.write(file_bytes)

        settings = load_settings(username)
        settings['avatar_url'] = f"/static/avatars/{filename}"
        save_settings(username, settings)

        log_activity("AVATAR_UPLOAD", {"filename": filename})
        return jsonify({"message": "Avatar updated", "avatar_url": settings['avatar_url']})


if __name__ == "__main__":
    _debug = os.getenv('FLASK_DEBUG', 'false').lower() == 'true'
    if _debug:
        print("[!] WARNING: Running in DEBUG mode — disable for production!")
    app.run(debug=_debug, host='0.0.0.0', port=5010, threaded=True)
